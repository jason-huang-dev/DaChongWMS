from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from django.core.files.uploadedfile import UploadedFile
from django.db.models.functions import Upper
from openpyxl import Workbook, load_workbook

from apps.inventory.models import InventoryBalance
from apps.organizations.models import Organization
from apps.products.models import Product

TEMPLATE_HEADERS = (
    "*Merchant SKU",
    "Product Name",
    "*Shelf",
    "*Available Stock",
    "Listing Time",
    "Actual Length",
    "Actual Width",
    "Actual Height",
    "Actual Weight",
    "*Unit of measurement (cm/g or in/lb)",
    "Merchant Code",
    "Customer Code",
)

ACCEPTED_MEASUREMENT_UNITS = {"cm/g", "in/lb"}
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class InventoryInformationImportRow:
    merchant_sku: str
    product_name: str
    shelf: str
    available_stock: str
    listing_time: str
    actual_length: str
    actual_width: str
    actual_height: str
    actual_weight: str
    measurement_unit: str
    merchant_code: str
    customer_code: str
    warehouse_name: str
    stock_status: str
    source: str

    def as_dict(self) -> dict[str, str]:
        return {
            "merchant_sku": self.merchant_sku,
            "product_name": self.product_name,
            "shelf": self.shelf,
            "available_stock": self.available_stock,
            "listing_time": self.listing_time,
            "actual_length": self.actual_length,
            "actual_width": self.actual_width,
            "actual_height": self.actual_height,
            "actual_weight": self.actual_weight,
            "measurement_unit": self.measurement_unit,
            "merchant_code": self.merchant_code,
            "customer_code": self.customer_code,
            "warehouse_name": self.warehouse_name,
            "stock_status": self.stock_status,
            "source": self.source,
        }


@dataclass(frozen=True)
class InventoryInformationImportResult:
    imported_rows: list[InventoryInformationImportRow]
    warnings: list[str]
    errors: list[str]

    def as_dict(self) -> dict[str, object]:
        return {
            "imported_rows": [row.as_dict() for row in self.imported_rows],
            "warnings": self.warnings,
            "errors": self.errors,
        }


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_sku(value: Any) -> str:
    return _normalize_text(value).upper()


def _normalize_shelf(value: Any) -> str:
    return _normalize_text(value).upper()


def _to_decimal(value: str) -> Decimal | None:
    if not value:
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError):
        return None


def _to_display_decimal(value: str) -> str:
    decimal_value = _to_decimal(value)
    if decimal_value is None:
        return value
    return format(decimal_value.normalize(), "f").rstrip("0").rstrip(".") or "0"


def _today_as_string(today: date | None = None) -> str:
    resolved_today = today or date.today()
    return resolved_today.isoformat()


def _normalize_listing_time(value: Any, fallback_value: str) -> str | None:
    normalized = _normalize_text(value)
    if not normalized:
        return fallback_value

    normalized = normalized.replace("/", "-")
    try:
        parsed = datetime.strptime(normalized, "%Y-%m-%d")
    except ValueError:
        return None
    return parsed.date().isoformat()


def _header_error_message() -> str:
    return f"Template header must exactly match: {' | '.join(TEMPLATE_HEADERS)}"


def build_inventory_information_template_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory Information"
    worksheet.append(list(TEMPLATE_HEADERS))
    worksheet.append(
        [
            "SKU-1001",
            "Bluetooth Scanner",
            "A-01-01",
            "120",
            "2026-03-29",
            "12.5",
            "8.2",
            "4.5",
            "260",
            "cm/g",
            "MER-001",
            "CUS-001",
        ]
    )

    for column_cells in worksheet.columns:
        max_length = max(len(_normalize_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max(max_length + 2, 18)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def parse_existing_inventory_information_rows(raw_value: str | None) -> list[dict[str, str]]:
    if not raw_value:
        return []

    try:
        parsed = json.loads(raw_value)
    except json.JSONDecodeError:
        return []

    if not isinstance(parsed, list):
        return []

    normalized_rows: list[dict[str, str]] = []
    for row in parsed:
        if not isinstance(row, dict):
            continue
        normalized_rows.append(
            {
                "merchant_sku": _normalize_sku(row.get("merchant_sku")),
                "shelf": _normalize_shelf(row.get("shelf")),
                "merchant_code": _normalize_text(row.get("merchant_code")),
                "customer_code": _normalize_text(row.get("customer_code")),
            }
        )
    return normalized_rows


def process_inventory_information_import(
    *,
    organization: Organization,
    workbook_file: UploadedFile,
    existing_rows: list[dict[str, str]] | None = None,
    import_date: date | None = None,
) -> InventoryInformationImportResult:
    file_name = _normalize_text(getattr(workbook_file, "name", ""))
    if not file_name.lower().endswith(".xlsx"):
        return InventoryInformationImportResult(
            imported_rows=[],
            warnings=[],
            errors=["Only .xlsx files are supported for inventory initialization imports."],
        )

    try:
        workbook = load_workbook(filename=BytesIO(workbook_file.read()), data_only=True, read_only=True)
    except Exception:
        return InventoryInformationImportResult(
            imported_rows=[],
            warnings=[],
            errors=["The uploaded file could not be read as a valid .xlsx workbook."],
        )

    worksheet = workbook[workbook.sheetnames[0]] if workbook.sheetnames else None
    if worksheet is None:
        return InventoryInformationImportResult(
            imported_rows=[],
            warnings=[],
            errors=[_header_error_message()],
        )

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return InventoryInformationImportResult(
            imported_rows=[],
            warnings=[],
            errors=[_header_error_message()],
        )

    header_row = [_normalize_text(value) for value in rows[0][: len(TEMPLATE_HEADERS)]]
    has_extra_header_cells = any(_normalize_text(value) for value in rows[0][len(TEMPLATE_HEADERS) :])
    if tuple(header_row) != TEMPLATE_HEADERS or has_extra_header_cells:
        return InventoryInformationImportResult(
            imported_rows=[],
            warnings=[],
            errors=[_header_error_message()],
        )

    data_rows = [
        row for row in rows[1:]
        if any(_normalize_text(value) for value in row)
    ]
    if not data_rows:
        return InventoryInformationImportResult(
            imported_rows=[],
            warnings=[],
            errors=["The template does not contain any inventory rows to import."],
        )

    if len(data_rows) > 5000:
        return InventoryInformationImportResult(
            imported_rows=[],
            warnings=[],
            errors=["The number of data rows imported at a time cannot exceed 5000."],
        )

    fallback_listing_time = _today_as_string(import_date)
    errors: list[str] = []
    warnings: list[str] = []
    parsed_rows: list[InventoryInformationImportRow] = []
    seen_input_sku_shelf: set[str] = set()

    for index, row in enumerate(data_rows, start=2):
        merchant_sku = _normalize_sku(row[0] if len(row) > 0 else "")
        product_name = _normalize_text(row[1] if len(row) > 1 else "")
        shelf = _normalize_shelf(row[2] if len(row) > 2 else "")
        available_stock_text = _normalize_text(row[3] if len(row) > 3 else "")
        listing_time = _normalize_listing_time(row[4] if len(row) > 4 else "", fallback_listing_time)
        actual_length = _normalize_text(row[5] if len(row) > 5 else "")
        actual_width = _normalize_text(row[6] if len(row) > 6 else "")
        actual_height = _normalize_text(row[7] if len(row) > 7 else "")
        actual_weight = _normalize_text(row[8] if len(row) > 8 else "")
        measurement_unit = _normalize_text(row[9] if len(row) > 9 else "").lower()
        merchant_code = _normalize_text(row[10] if len(row) > 10 else "")
        customer_code = _normalize_text(row[11] if len(row) > 11 else "")

        if merchant_sku and shelf:
            sku_shelf_key = f"{merchant_sku}::{shelf}"
            if sku_shelf_key in seen_input_sku_shelf:
                warnings.append(
                    f"Duplicate Merchant SKU + Shelf detected for {merchant_sku} at {shelf}. The first row was kept."
                )
                continue
            seen_input_sku_shelf.add(sku_shelf_key)

        if not merchant_sku:
            errors.append(f"Row {index}: Merchant SKU is required.")
        if not shelf:
            errors.append(f"Row {index}: Shelf is required.")

        available_stock = _to_decimal(available_stock_text)
        if not available_stock_text or available_stock is None or available_stock < 0:
            errors.append(f"Row {index}: Available Stock must be a valid non-negative number.")

        if listing_time is None:
            errors.append(f"Row {index}: Listing Time must use YYYY-MM-DD or YYYY/MM/DD.")

        if not all((actual_length, actual_width, actual_height, actual_weight)):
            errors.append(
                f'Row {index}: Actual Length, Width, Height, and Weight are required when "Product Management" has no size profile.'
            )

        for label, raw_value in (
            ("Actual Length", actual_length),
            ("Actual Width", actual_width),
            ("Actual Height", actual_height),
            ("Actual Weight", actual_weight),
        ):
            if not raw_value:
                continue
            numeric_value = _to_decimal(raw_value)
            if numeric_value is None or numeric_value <= 0:
                errors.append(f"Row {index}: {label} must be a positive number.")

        if measurement_unit not in ACCEPTED_MEASUREMENT_UNITS:
            errors.append(f"Row {index}: Unit of measurement must be either cm/g or in/lb.")

        parsed_rows.append(
            InventoryInformationImportRow(
                merchant_sku=merchant_sku,
                product_name=product_name or merchant_sku,
                shelf=shelf,
                available_stock=_to_display_decimal(available_stock_text),
                listing_time=listing_time or fallback_listing_time,
                actual_length=_to_display_decimal(actual_length),
                actual_width=_to_display_decimal(actual_width),
                actual_height=_to_display_decimal(actual_height),
                actual_weight=_to_display_decimal(actual_weight),
                measurement_unit=measurement_unit,
                merchant_code=merchant_code,
                customer_code=customer_code,
                warehouse_name="",
                stock_status="AVAILABLE",
                source="imported",
            )
        )

    if errors:
        return InventoryInformationImportResult(imported_rows=[], warnings=[], errors=sorted(set(errors)))

    existing_rows = existing_rows or []
    normalized_uploaded_skus = [_normalize_sku(row.merchant_sku) for row in parsed_rows if _normalize_sku(row.merchant_sku)]
    existing_sku_set = {
        _normalize_sku(row["merchant_sku"])
        for row in existing_rows
        if _normalize_sku(row.get("merchant_sku"))
    }
    existing_sku_set.update(
        _normalize_sku(sku)
        for sku in Product.objects.filter(organization=organization)
        .annotate(normalized_sku=Upper("sku"))
        .filter(normalized_sku__in=normalized_uploaded_skus)
        .values_list("normalized_sku", flat=True)
    )
    existing_sku_set.update(
        _normalize_sku(sku)
        for sku in InventoryBalance.objects.filter(organization=organization)
        .annotate(normalized_sku=Upper("product__sku"))
        .filter(normalized_sku__in=normalized_uploaded_skus)
        .values_list("normalized_sku", flat=True)
    )

    for row in parsed_rows:
        if row.merchant_sku in existing_sku_set:
            errors.append(f"{row.merchant_sku} already exists in the Inventory List, so initialization failed.")

    rows_by_sku: dict[str, list[InventoryInformationImportRow]] = {}
    for row in parsed_rows:
        rows_by_sku.setdefault(row.merchant_sku, []).append(row)

    for merchant_sku, merchant_rows in rows_by_sku.items():
        if len(merchant_rows) > 1:
            errors.append(f"{merchant_sku} appears multiple times. Inventory initialization allows only one row per Merchant SKU.")

    return InventoryInformationImportResult(
        imported_rows=[] if errors else sorted(parsed_rows, key=lambda row: (row.source != "imported", row.merchant_sku, row.shelf)),
        warnings=[] if errors else sorted(set(warnings)),
        errors=sorted(set(errors)),
    )
