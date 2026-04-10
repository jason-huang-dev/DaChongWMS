from __future__ import annotations

from datetime import date
from io import BytesIO
from decimal import Decimal

from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.inbound.services.inbound_service import (
    CreatePurchaseOrderInput,
    CreatePurchaseOrderLineInput,
    CreateReceiptInput,
    ReceiptLineInput,
    create_purchase_order,
    record_receipt,
)
from apps.inventory.models import InventoryBalance, InventoryStatus
from apps.iam.models import Role
from apps.locations.services.location_service import (
    CreateLocationInput,
    CreateLocationTypeInput,
    CreateZoneInput,
    create_location,
    create_location_type,
    create_zone,
)
from apps.organizations.tests.test_factories import (
    add_membership,
    assign_role,
    grant_role_permission,
    make_customer_account,
    make_organization,
    make_role,
    make_user,
)
from apps.products.models import Product
from apps.warehouse.models import Warehouse


class InventoryAPITests(TestCase):
    def setUp(self) -> None:
        self.client = APIClient()
        self.organization = make_organization()
        self.manager = make_user("manager@example.com")
        self.viewer = make_user("viewer@example.com")
        self.manager_membership = add_membership(self.manager, self.organization)
        self.viewer_membership = add_membership(self.viewer, self.organization)
        self.warehouse = Warehouse.objects.create(
            organization=self.organization,
            name="Primary",
            code="WH-A",
        )
        self.customer_account = make_customer_account(
            self.organization,
            name="Retail Client",
            code="RTL-1",
        )
        self.product = Product.objects.create(
            organization=self.organization,
            sku="SKU-001",
            name="Scanner",
        )
        zone = create_zone(
            CreateZoneInput(
                organization=self.organization,
                warehouse=self.warehouse,
                code="STOR",
                name="Storage",
            )
        )
        location_type = create_location_type(
            CreateLocationTypeInput(
                organization=self.organization,
                code="BULK",
                name="Bulk",
            )
        )
        self.location = create_location(
            CreateLocationInput(
                organization=self.organization,
                warehouse=self.warehouse,
                zone=zone,
                location_type=location_type,
                code="A-01-01",
            )
        )

        manager_role = make_role(Role.SystemCode.MANAGER)
        staff_role = make_role(Role.SystemCode.STAFF)

        for codename in ("view_inventory", "manage_inventory_records", "manage_inventory_configuration"):
            permission = Permission.objects.get(
                content_type__app_label="inventory",
                codename=codename,
            )
            grant_role_permission(manager_role, permission)

        view_permission = Permission.objects.get(
            content_type__app_label="inventory",
            codename="view_inventory",
        )
        grant_role_permission(staff_role, view_permission)

        assign_role(self.manager_membership, manager_role)
        assign_role(self.viewer_membership, staff_role)

    def _build_import_file(self, rows: list[list[object]], file_name: str = "inventory.xlsx") -> SimpleUploadedFile:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Inventory Information"
        for row in rows:
            worksheet.append(row)

        stream = BytesIO()
        workbook.save(stream)
        return SimpleUploadedFile(
            file_name,
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_manager_can_record_inventory_and_configure_adjustments(self) -> None:
        self.client.force_authenticate(self.manager)

        movement_response = self.client.post(
            reverse(
                "organization-inventory-movement-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {
                "warehouse_id": self.warehouse.id,
                "product_id": self.product.id,
                "to_location_id": self.location.id,
                "movement_type": "RECEIPT",
                "stock_status": "AVAILABLE",
                "quantity": "8.0000",
                "unit_cost": "3.2500",
                "currency": "usd",
                "reason": "Inbound receipt",
            },
            format="json",
        )
        self.assertEqual(movement_response.status_code, status.HTTP_201_CREATED)

        balances_response = self.client.get(
            reverse(
                "organization-inventory-balance-list",
                kwargs={"organization_id": self.organization.id},
            )
        )
        self.assertEqual(balances_response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(balances_response.data), 1)

        reason_response = self.client.post(
            reverse(
                "organization-inventory-adjustment-reason-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {
                "code": "damage",
                "name": "Damage",
                "description": "Damage write-off",
                "direction": "DECREASE",
                "requires_approval": True,
                "is_active": True,
            },
            format="json",
        )
        self.assertEqual(reason_response.status_code, status.HTTP_201_CREATED)

    def test_manager_can_record_atomic_inventory_adjustment_list(self) -> None:
        self.client.force_authenticate(self.manager)
        product_two = Product.objects.create(
            organization=self.organization,
            sku="SKU-002",
            name="Label Printer",
        )

        first_receipt_response = self.client.post(
            reverse(
                "organization-inventory-movement-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {
                "warehouse_id": self.warehouse.id,
                "product_id": self.product.id,
                "to_location_id": self.location.id,
                "movement_type": "RECEIPT",
                "stock_status": "AVAILABLE",
                "quantity": "8.0000",
                "unit_cost": "3.2500",
                "reason": "Initial receipt",
            },
            format="json",
        )
        self.assertEqual(first_receipt_response.status_code, status.HTTP_201_CREATED)

        second_receipt_response = self.client.post(
            reverse(
                "organization-inventory-movement-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {
                "warehouse_id": self.warehouse.id,
                "product_id": product_two.id,
                "to_location_id": self.location.id,
                "movement_type": "RECEIPT",
                "stock_status": "AVAILABLE",
                "quantity": "5.0000",
                "unit_cost": "4.5000",
                "reason": "Initial receipt",
            },
            format="json",
        )
        self.assertEqual(second_receipt_response.status_code, status.HTTP_201_CREATED)

        balances = InventoryBalance.objects.select_related("product").filter(organization=self.organization).order_by("product__sku")
        self.assertEqual(balances.count(), 2)
        first_balance, second_balance = list(balances)

        adjustment_response = self.client.post(
            reverse(
                "organization-inventory-movement-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {
                "warehouse_id": self.warehouse.id,
                "adjustment_type": "Good Product Adjustment",
                "note": "Cycle count review",
                "items": [
                    {
                        "balance_id": first_balance.id,
                        "movement_type": "ADJUSTMENT_OUT",
                        "quantity": "2.0000",
                    },
                    {
                        "balance_id": second_balance.id,
                        "movement_type": "ADJUSTMENT_IN",
                        "quantity": "3.0000",
                    },
                ],
            },
            format="json",
        )

        self.assertEqual(adjustment_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(adjustment_response.data["count"], 2)
        self.assertEqual(adjustment_response.data["warehouse_id"], self.warehouse.id)
        self.assertEqual(
            adjustment_response.data["reason"],
            "Good Product Adjustment: Cycle count review",
        )
        self.assertTrue(adjustment_response.data["reference_code"].startswith("ADJ-"))
        self.assertEqual(len(adjustment_response.data["results"]), 2)
        self.assertEqual(
            {result["reference_code"] for result in adjustment_response.data["results"]},
            {adjustment_response.data["reference_code"]},
        )

        first_balance.refresh_from_db()
        second_balance.refresh_from_db()
        self.assertEqual(first_balance.on_hand_qty, Decimal("6.0000"))
        self.assertEqual(second_balance.on_hand_qty, Decimal("8.0000"))

    def test_viewer_can_list_balances_but_cannot_record_inventory(self) -> None:
        self.client.force_authenticate(self.viewer)

        list_response = self.client.get(
            reverse(
                "organization-inventory-balance-list",
                kwargs={"organization_id": self.organization.id},
            )
        )
        create_response = self.client.post(
            reverse(
                "organization-inventory-movement-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {
                "warehouse_id": self.warehouse.id,
                "product_id": self.product.id,
                "to_location_id": self.location.id,
                "movement_type": "RECEIPT",
                "stock_status": "AVAILABLE",
                "quantity": "1.0000",
            },
            format="json",
        )

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(create_response.status_code, status.HTTP_403_FORBIDDEN)

    def test_manager_can_list_backend_driven_inventory_movements(self) -> None:
        self.client.force_authenticate(self.manager)
        purchase_order = create_purchase_order(
            CreatePurchaseOrderInput(
                organization=self.organization,
                warehouse=self.warehouse,
                customer_account=self.customer_account,
                po_number="PO-2001",
                line_items=(
                    CreatePurchaseOrderLineInput(
                        line_number=1,
                        product=self.product,
                        ordered_qty=Decimal("8.0000"),
                        unit_cost=Decimal("3.2500"),
                    ),
                ),
            )
        )
        purchase_order_line = purchase_order.lines.get(line_number=1)
        receipt = record_receipt(
            payload=CreateReceiptInput(
                purchase_order=purchase_order,
                warehouse=self.warehouse,
                receipt_location=self.location,
                receipt_number="RCPT-2001",
                reference_code="R5ZG260402210815",
                line_items=(
                    ReceiptLineInput(
                        purchase_order_line=purchase_order_line,
                        received_qty=Decimal("8.0000"),
                        stock_status=InventoryStatus.AVAILABLE,
                        lot_number="BO20260402008",
                        serial_number="R5ZG260402210815",
                        unit_cost=Decimal("3.2500"),
                    ),
                ),
            ),
            operator_name="manager@example.com",
        )

        list_response = self.client.get(
            reverse(
                "organization-inventory-movement-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {
                "warehouse_id": self.warehouse.id,
                "page_size": 50,
                "movementTypes": '["RECEIPT"]',
                "merchantSku": "SKU-001",
                "locationCode": "A-01-01",
                "performedBy": "manager@example.com",
                "referenceCode": receipt.receipt_number,
                "matchMode": "exact",
            },
        )

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(list_response.data["count"], 1)
        row = list_response.data["results"][0]
        self.assertEqual(row["merchantSku"], "SKU-001")
        self.assertEqual(row["productName"], "Scanner")
        self.assertEqual(row["warehouseName"], "Primary")
        self.assertEqual(row["movementType"], "RECEIPT")
        self.assertEqual(row["performedBy"], "manager@example.com")
        self.assertEqual(row["resultingLocationCode"], "A-01-01")
        self.assertEqual(row["clientCode"], "RTL-1")
        self.assertEqual(row["clientName"], "Retail Client")
        self.assertEqual(row["entryTypeLabel"], "Standard Stock-in")
        self.assertEqual(row["purchaseOrderNumber"], "PO-2001")
        self.assertEqual(row["receiptNumber"], "RCPT-2001")
        self.assertEqual(row["sourceDocumentNumber"], "RCPT-2001")
        self.assertEqual(row["batchNumber"], "BO20260402008")
        self.assertEqual(row["serialNumber"], "R5ZG260402210815")
        self.assertEqual(
            row["linkedDocumentNumbers"],
            [
                {"label": "Stock-in No.", "value": "RCPT-2001"},
                {"label": "Receiving Serial Number", "value": "R5ZG260402210815"},
                {"label": "Listing Serial Number", "value": "PT-RCPT-2001-1"},
            ],
        )
        self.assertEqual(
            row["sourceDocumentNumbers"],
            [
                {"label": "Purchase Order", "value": "PO-2001"},
                {"label": "Reference", "value": "R5ZG260402210815"},
            ],
        )
        self.assertEqual(row["quantity"], 8)
        self.assertEqual(
            list_response.data["filterOptions"]["warehouses"],
            [{"value": str(self.warehouse.id), "label": "Primary"}],
        )
        self.assertEqual(
            list_response.data["filterOptions"]["movementTypes"],
            [{"value": "RECEIPT", "label": "Receipt"}],
        )

    def test_manager_can_list_backend_driven_inventory_information(self) -> None:
        self.client.force_authenticate(self.manager)
        InventoryBalance.objects.create(
            organization=self.organization,
            warehouse=self.warehouse,
            location=self.location,
            product=self.product,
            stock_status=InventoryStatus.AVAILABLE,
            on_hand_qty="8.0000",
            allocated_qty="2.0000",
            hold_qty="1.0000",
            unit_cost="3.2500",
            currency="USD",
        )

        response = self.client.get(
            reverse(
                "organization-inventory-information-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {"warehouse_id": self.warehouse.id, "page_size": 50},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        row = response.data["results"][0]
        self.assertEqual(row["merchantSku"], "SKU-001")
        self.assertEqual(row["productName"], "Scanner")
        self.assertEqual(row["warehouseName"], "Primary")
        self.assertEqual(row["availableStock"], 5)
        self.assertEqual(row["totalInventory"], 8)
        self.assertEqual(
            response.data["filterOptions"]["warehouses"],
            [{"value": "Primary", "label": "Primary"}],
        )
        self.assertEqual(
            response.data["filterOptions"]["skus"],
            [{"value": "SKU-001", "label": "SKU-001"}],
        )

    def test_manager_can_hide_zero_stock_inventory_information_rows(self) -> None:
        self.client.force_authenticate(self.manager)
        zero_stock_product = Product.objects.create(
            organization=self.organization,
            sku="SKU-002",
            name="Empty Bin Scanner",
        )
        InventoryBalance.objects.create(
            organization=self.organization,
            warehouse=self.warehouse,
            location=self.location,
            product=self.product,
            stock_status=InventoryStatus.AVAILABLE,
            on_hand_qty="8.0000",
            allocated_qty="2.0000",
            hold_qty="1.0000",
            unit_cost="3.2500",
            currency="USD",
        )
        InventoryBalance.objects.create(
            organization=self.organization,
            warehouse=self.warehouse,
            location=self.location,
            product=zero_stock_product,
            stock_status=InventoryStatus.AVAILABLE,
            on_hand_qty="0.0000",
            allocated_qty="0.0000",
            hold_qty="0.0000",
            unit_cost="3.2500",
            currency="USD",
        )

        response = self.client.get(
            reverse(
                "organization-inventory-information-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {"warehouse_id": self.warehouse.id, "hideZeroStock": "true", "page_size": 50},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual([row["merchantSku"] for row in response.data["results"]], ["SKU-001"])

    def test_manager_can_download_inventory_information_template(self) -> None:
        self.client.force_authenticate(self.manager)

        response = self.client.get(
            reverse(
                "organization-inventory-information-import-template",
                kwargs={"organization_id": self.organization.id},
            )
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        header = [cell for cell in next(worksheet.iter_rows(values_only=True))]
        self.assertEqual(
            header,
            [
                "*Merchant SKU",
                "Product Name",
                "*Shelf",
                "*Available Stock",
                "Listing Time",
                "Actual Length",
                "Actual Width",
                "Actual Height",
                "Actual Weight",
                "*Unit of measurement (cm/g or in/lb)",
                "Merchant Code",
                "Customer Code",
            ],
        )

    def test_manager_can_upload_inventory_information_workbook_for_backend_processing(self) -> None:
        self.client.force_authenticate(self.manager)

        upload_file = self._build_import_file(
            [
                [
                    "*Merchant SKU",
                    "Product Name",
                    "*Shelf",
                    "*Available Stock",
                    "Listing Time",
                    "Actual Length",
                    "Actual Width",
                    "Actual Height",
                    "Actual Weight",
                    "*Unit of measurement (cm/g or in/lb)",
                    "Merchant Code",
                    "Customer Code",
                ],
                ["SKU-NEW-001", "New Scanner", "B-01-01", "12", "", "12.5", "8.2", "4.5", "260", "cm/g", "MER-001", "CUS-001"],
                ["SKU-NEW-001", "New Scanner", "B-01-01", "99", "2026-03-29", "12.5", "8.2", "4.5", "260", "cm/g", "MER-001", "CUS-001"],
            ]
        )

        response = self.client.post(
            reverse(
                "organization-inventory-information-import-upload",
                kwargs={"organization_id": self.organization.id},
            ),
            {"file": upload_file, "warehouse_id": str(self.warehouse.id)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["errors"], [])
        self.assertEqual(
            response.data["warnings"],
            ["Duplicate Merchant SKU + Shelf detected for SKU-NEW-001 at B-01-01. The first row was kept."],
        )
        self.assertEqual(len(response.data["imported_rows"]), 1)
        self.assertEqual(response.data["imported_rows"][0]["merchant_sku"], "SKU-NEW-001")
        self.assertEqual(response.data["imported_rows"][0]["listing_time"], date.today().isoformat())

        list_response = self.client.get(
            reverse(
                "organization-inventory-information-list",
                kwargs={"organization_id": self.organization.id},
            ),
            {"warehouse_id": self.warehouse.id, "page_size": 50},
        )

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(list_response.data["count"], 1)
        self.assertEqual(list_response.data["results"][0]["merchantSku"], "SKU-NEW-001")
        self.assertEqual(list_response.data["results"][0]["source"], "imported")

    def test_manager_inventory_import_fails_when_product_already_exists(self) -> None:
        self.client.force_authenticate(self.manager)

        upload_file = self._build_import_file(
            [
                [
                    "*Merchant SKU",
                    "Product Name",
                    "*Shelf",
                    "*Available Stock",
                    "Listing Time",
                    "Actual Length",
                    "Actual Width",
                    "Actual Height",
                    "Actual Weight",
                    "*Unit of measurement (cm/g or in/lb)",
                    "Merchant Code",
                    "Customer Code",
                ],
                ["SKU-001", "Scanner", "B-01-01", "12", "2026-03-29", "12.5", "8.2", "4.5", "260", "cm/g", "MER-001", "CUS-001"],
            ]
        )

        response = self.client.post(
            reverse(
                "organization-inventory-information-import-upload",
                kwargs={"organization_id": self.organization.id},
            ),
            {"file": upload_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["imported_rows"], [])
        self.assertEqual(response.data["warnings"], [])
        self.assertEqual(
            response.data["errors"],
            ["SKU-001 already exists in the Inventory List, so initialization failed."],
        )

    def test_manager_inventory_import_fails_when_existing_product_matches_case_insensitively(self) -> None:
        self.client.force_authenticate(self.manager)
        Product.objects.create(
            organization=self.organization,
            sku="sku-case-001",
            name="Case Variant Scanner",
        )

        upload_file = self._build_import_file(
            [
                [
                    "*Merchant SKU",
                    "Product Name",
                    "*Shelf",
                    "*Available Stock",
                    "Listing Time",
                    "Actual Length",
                    "Actual Width",
                    "Actual Height",
                    "Actual Weight",
                    "*Unit of measurement (cm/g or in/lb)",
                    "Merchant Code",
                    "Customer Code",
                ],
                ["SKU-CASE-001", "Scanner", "B-01-01", "12", "2026-03-29", "12.5", "8.2", "4.5", "260", "cm/g", "MER-001", "CUS-001"],
            ]
        )

        response = self.client.post(
            reverse(
                "organization-inventory-information-import-upload",
                kwargs={"organization_id": self.organization.id},
            ),
            {"file": upload_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["imported_rows"], [])
        self.assertEqual(response.data["warnings"], [])
        self.assertEqual(
            response.data["errors"],
            ["SKU-CASE-001 already exists in the Inventory List, so initialization failed."],
        )
